#!/usr/bin/env python3
import argparse, logging
from pathlib import Path
from openpyxl import load_workbook
from cfg import PATHS, AI_CFG

LOG_PATH = Path(PATHS.get("logs","./logs")) / "migration_artifacts.log"
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(filename=str(LOG_PATH), level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")

def s(v): return "" if v is None else str(v).strip()
def up(v): return s(v).upper()

def find_col(headers, keyword):
    for i,h in enumerate(headers):
        if keyword.lower() in s(h).lower().replace("_"," "):
            return i
    return None

def generate_artifacts(xlsx:Path, out_dir:Path):
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "baseline" in name.lower():
            ws = wb[name]; break
    if ws is None: ws = wb.active

    # ✅ FIXED: header row now row 1
    headers = [s(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    i_status = find_col(headers, "validation")
    if i_status is None:
        print("❌ Could not find Validation_Status column. Last headers:", headers[-10:])
        return

    print(f"✅ Header row detected (len={len(headers)}). Validation col index = {i_status}")

    out_dir.mkdir(parents=True, exist_ok=True)
    ddl_file = out_dir / "generated_ddl.sql"

    with ddl_file.open("w") as f:
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row): continue
            status = up(s(row[i_status])) if i_status < len(row) else ""
            if status != "PASS": continue
            dest_table = s(row[find_col(headers,"destination table")])
            dest_col = s(row[find_col(headers,"destination column name")])
            dtype = s(row[find_col(headers,"data type")]) or "VARCHAR"
            if not dest_table or not dest_col: continue
            f.write(f"ALTER TABLE {dest_table} ADD {dest_col} {dtype};\n")

    print(f"[OK] DDL → {ddl_file}")
    wb.close()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--validated-xlsx", required=True)
    ap.add_argument("--out-dir", required=True)
    args = ap.parse_args()
    generate_artifacts(Path(args.validated_xlsx), Path(args.out_dir))

if __name__ == "__main__":
    main()
