#!/usr/bin/env python3
"""
validate_dmw_vs_ddl_stream.py
- Deterministic 6-rule validator
- DDL checks
- Table Details cross-check
- Missing_Tables_Report sheet
- Summary_Stats sheet
- Header trim so appended columns sit right after last real header
- Optional AI queue generation via config (no inference in main run)
"""

import re, json, yaml, argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook, Workbook

# ---------- Helpers ----------
def s(x): return "" if x is None else str(x).strip()
def up(x): return s(x).upper()
def canon_yesno(v):
    v = up(v)
    if v in ("Y","YES","TRUE","1"): return "YES"
    if v in ("N","NO","FALSE","0"): return "NO"
    return v

# ---------- DDL Parser ----------
def parse_type_sizes(t):
    t = s(t)
    m = re.search(r"\(([^)]+)\)", t)
    if not m: return ("","","")
    parts = [p.strip() for p in m.group(1).split(",")]
    if len(parts)==1: return (parts[0],"","")
    if len(parts)>=2: return ("",parts[0],parts[1])
    return ("","","")

def strip_quotes(v):
    v = s(v)
    if len(v)>=2 and v[0] in "'\"" and v[-1] in "'\"": return v[1:-1]
    return v

def canon_type(t):
    t = up(t)
    base = re.sub(r"\s*\(.*\)","",t).strip()
    aliases = {
        "VARCHAR2":"VARCHAR","NVARCHAR2":"NVARCHAR","NUMBER":"DECIMAL","NUMERIC":"DECIMAL",
        "INT":"INTEGER","INT4":"INTEGER","INT8":"BIGINT",
        "SMALLDATETIME":"DATETIME","DATETIME2":"DATETIME","DATETIMEOFFSET":"DATETIME",
        "TIMESTAMP WITH TIME ZONE":"TIMESTAMP","TIMESTAMP WITHOUT TIME ZONE":"TIMESTAMP",
        "BOOL":"BOOLEAN","BIT":"BOOLEAN","DOUBLE PRECISION":"DOUBLE"
    }
    return aliases.get(base, base)

def split_create_table(sql):
    out=[]; i=0; n=len(sql)
    pat=re.compile(r"CREATE\s+TABLE\s+(.+?)\s*\(", re.IGNORECASE|re.DOTALL)
    while True:
        m=pat.search(sql,i)
        if not m: break
        depth=0; j=m.end()-1; found=False
        while j<n:
            ch=sql[j]
            if ch=='(': depth+=1
            elif ch==')':
                depth-=1
                if depth<=0:
                    k=j+1
                    while k<n and sql[k].isspace(): k+=1
                    if k<n and sql[k]==';':
                        out.append(sql[m.start():k+1]); i=k+1; found=True; break
            j+=1
        if not found: i=m.end()
    return out

def parse_create_table(stmt):
    m=re.search(r"CREATE\s+TABLE\s+(.+?)\s*\(", stmt, re.IGNORECASE|re.DOTALL)
    if not m: return None
    table=up(re.sub(r'[\[\]"]','',m.group(1).split('.')[-1].strip()))
    body_start=m.end(); depth=1; i=body_start; n=len(stmt)
    token=[]
    while i<n and depth>0:
        ch=stmt[i]; token.append(ch)
        if ch=='(': depth+=1
        elif ch==')': depth-=1
        i+=1
    body=''.join(token[:-1])
    cols={}; pk=set(); uq=set()
    for line in re.split(r',(?=(?:[^()]|\([^()]*\))*$)',body):
        l=line.strip()
        if not l: continue
        if l.upper().startswith("CONSTRAINT"):  # ignore table constraints in this lightweight parser
            continue
        mcol=re.match(r'^("?[A-Za-z0-9_]+"?)\s+(.+)$',l)
        if not mcol: continue
        col=up(mcol.group(1).strip('"[]'))
        rest=mcol.group(2)
        dtype=re.split(r'\s+',rest)[0]
        char_len,prec,sca=parse_type_sizes(dtype)
        not_null="NOT NULL" in up(rest)
        default_m=re.search(r"DEFAULT\s+([^ ]+)",rest,re.IGNORECASE)
        default=strip_quotes(default_m.group(1)) if default_m else ""
        cols[col]={
            "data_type":dtype,
            "char_length":s(char_len),
            "precision":s(prec),
            "scale":s(sca),
            "is_nullable_yesno":"NO" if not_null else "YES",
            "default":default
        }
    return table,cols,pk,uq

def build_ddl_index(sql):
    ddl={}; pk=defaultdict(set); uq=defaultdict(set)
    for stmt in split_create_table(sql):
        p=parse_create_table(stmt)
        if p:
            t,c,pkc,uqc=p
            ddl[t]=c; pk[t]|=pkc; uq[t]|=uqc
    return ddl,pk,uq

# ---------- Table Details reader ----------
def load_table_details_tables(xlsx_path):
    tables=set()
    try:
        wb=load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return tables
    candidates=[n for n in wb.sheetnames if "table" in n.lower()]
    for sh in candidates:
        ws=wb[sh]
        header_row=1
        for i, r in enumerate(ws.iter_rows(min_row=1,max_row=10,values_only=True), start=1):
            vals=[up(s(v)) for v in r]
            if any(v in ("TABLE NAME","TABLE","DESTINATION TABLE") for v in vals):
                header_row=i; break
        headers=[s(c) for c in next(ws.iter_rows(min_row=header_row,max_row=header_row,values_only=True))]
        tcol=None
        for idx,h in enumerate(headers):
            if up(h) in ("TABLE NAME","TABLE","DESTINATION TABLE"):
                tcol=idx+1; break
        if not tcol: continue
        empty=0
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            name=s(row[tcol-1] if tcol-1 < len(row) else "")
            if name:
                tables.add(up(name)); empty=0
            else:
                empty+=1
                if empty>200: break
    wb.close()
    return tables

# ---------- Main ----------
def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--dmw-xlsx",required=True)
    ap.add_argument("--ddl-sql",required=True)
    ap.add_argument("--out",required=True)
    ap.add_argument("--sheet",default="Baseline Data Model")
    ap.add_argument("--max-rows",type=int,default=10000)
    ap.add_argument("--config",help="Optional YAML/JSON config to control AI")
    args=ap.parse_args()

    # Config
    config={}
    if args.config and Path(args.config).exists():
        try:
            if args.config.endswith((".yaml",".yml")):
                config=yaml.safe_load(Path(args.config).read_text())
            else:
                config=json.loads(Path(args.config).read_text())
            print(f"[INFO] Loaded config from {args.config}")
        except Exception as e:
            print(f"[WARN] Failed to read config {args.config}: {e}")
    ai_enabled = bool(config.get("ai",{}).get("enabled", False))
    print(f"[INFO] AI Enabled: {ai_enabled}")

    # Load DDL
    ddl_text=Path(args.ddl_sql).read_text(encoding="utf-8",errors="ignore")
    ddl,pk,uq=build_ddl_index(ddl_text)
    ddl_tables=set(ddl.keys())

    # Load Table Details set
    table_details = load_table_details_tables(args.dmw_xlsx)

    # Load DMW sheet
    wb=load_workbook(args.dmw_xlsx,read_only=True,data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"[ERROR] Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
    ws=wb[args.sheet]

    header_row=2
    headers=[s(c) for c in next(ws.iter_rows(min_row=header_row,max_row=header_row,values_only=True))]
    while headers and (headers[-1]=="" or headers[-1].startswith("Unnamed")):
        headers.pop()
    print(f"[INFO] Trimmed header to {len(headers)} columns (last: {headers[-1] if headers else 'N/A'})")

    # Output workbook
    out_wb=Workbook(write_only=True)
    out_ws=out_wb.create_sheet("Baseline Data Model_output")

    appended_headers=[
        "Mapping_Check","ChangeTracking_Check","TableConsistency_Check",
        "DDLAlignment_Check","ReferenceSubset_Check","VersionDiff_Check",
        "Validation_Status","Validation_Remarks","Change Log (for data migration reference)"
    ]
    out_ws.append(headers+appended_headers)

    def col_idx(name):
        try: return headers.index(name)
        except ValueError: return None

    ix_dest_table=col_idx("Destination Table")
    ix_dest_col=col_idx("Destination Column Name")
    ix_dtype=col_idx("Data Type")
    ix_len=col_idx("Max Length")
    ix_null=col_idx("Is it Nullable? Yes/No")
    ix_trans=col_idx("Transformation Description") or col_idx("Transformation Description (Transformation Logic)")
    ix_migr=col_idx("Migrating or Not (Yes/No)")
    ix_intro=col_idx("Introduced Sprint")
    ix_lastupd=col_idx("Last Updated in Sprint")
    ix_reason=col_idx("Reason for Not Migrating")

    # Safety cap
    max_row=ws.max_row
    if max_row>args.max_rows:
        print(f"[WARN] Sheet indicates {max_row} rows — likely ghost rows; limiting to {args.max_rows}.")
        max_row=args.max_rows

    # Stats + tracking
    rules=["Mapping_Check","ChangeTracking_Check","TableConsistency_Check",
           "DDLAlignment_Check","ReferenceSubset_Check","VersionDiff_Check"]
    stats={r:{"PASS":0,"FAIL":0,"SKIPPED":0} for r in rules}
    seen_dest_tables=set()
    ai_candidates=[] if ai_enabled else None

    # Iterate rows
    start_row = header_row+1
    for r_idx, row in enumerate(ws.iter_rows(min_row=start_row,max_row=max_row,values_only=True), start=start_row):
        vals=[s(c) for c in row[:len(headers)]]
        dest_table=up(vals[ix_dest_table]) if ix_dest_table is not None and ix_dest_table < len(vals) else ""
        dest_col=up(vals[ix_dest_col]) if ix_dest_col is not None and ix_dest_col < len(vals) else ""

        # Rule 1: Mapping completeness
        mig=canon_yesno(vals[ix_migr]) if ix_migr is not None and ix_migr < len(vals) else ""
        mapping_ok=True; mapping_msg=[]
        if mig=="YES":
            need = {
                "Data Type": vals[ix_dtype] if ix_dtype is not None and ix_dtype < len(vals) else "",
                "Max Length": vals[ix_len] if ix_len is not None and ix_len < len(vals) else "",
                "Is it Nullable? Yes/No": vals[ix_null] if ix_null is not None and ix_null < len(vals) else "",
                "Transformation Description": vals[ix_trans] if ix_trans is not None and ix_trans < len(vals) else "",
            }
            missing=[k for k,v in need.items() if s(v)==""]
            if missing:
                mapping_ok=False; mapping_msg.append("Missing: "+", ".join(missing))
        if mig=="NO":
            reason = vals[ix_reason] if ix_reason is not None and ix_reason < len(vals) else ""
            if s(reason)=="":
                mapping_ok=False; mapping_msg.append("Missing Reason for Not Migrating")
        mapping_check="PASS" if mapping_ok else "FAIL"
        stats["Mapping_Check"][mapping_check]+=1 if mapping_check in ("PASS","FAIL") else 0

        # Rule 2: Change tracking
        intro=s(vals[ix_intro]) if ix_intro is not None and ix_intro < len(vals) else ""
        lu=s(vals[ix_lastupd]) if ix_lastupd is not None and ix_lastupd < len(vals) else ""
        change_log=""
        ct_ok=True
        if not intro or not lu:
            ct_ok=False
        if intro and lu and intro!=lu:
            change_log=f"Changed in Sprint {lu}"
        change_check="PASS" if ct_ok else "FAIL"
        stats["ChangeTracking_Check"][change_check]+=1 if change_check in ("PASS","FAIL") else 0

        # Rule 3: Table consistency (Table Details)
        if dest_table:
            seen_dest_tables.add(dest_table)
        if table_details:
            table_consistency = "PASS" if (dest_table in table_details or dest_table=="") else "FAIL"
        else:
            table_consistency = "SKIPPED"
        stats["TableConsistency_Check"][table_consistency]+=1 if table_consistency in ("PASS","FAIL","SKIPPED") else 0

        # Rule 4: DDL alignment
        ddl_align="PASS"
        remark=[]
        if dest_table and dest_table not in ddl:
            ddl_align="FAIL"; remark.append("Table not in DDL")
        elif dest_table and dest_col and dest_col not in ddl.get(dest_table, {}):
            ddl_align="FAIL"; remark.append("Column not in DDL")
        stats["DDLAlignment_Check"][ddl_align]+=1 if ddl_align in ("PASS","FAIL") else 0

        # Rule 5: Reference subset (heuristic)
        ref_check="SKIPPED"
        if "_REF" in dest_table:
            ref_check="PASS" if (vals[ix_dtype] if ix_dtype is not None and ix_dtype < len(vals) else "") else "FAIL"
        stats["ReferenceSubset_Check"][ref_check]+=1 if ref_check in ("PASS","FAIL","SKIPPED") else 0

        # Rule 6: Version diff (placeholder)
        ver_check="SKIPPED"
        stats["VersionDiff_Check"][ver_check]+=1

        status="PASS" if all(x=="PASS" for x in [
            mapping_check,
            change_check,
            (table_consistency if table_consistency!="SKIPPED" else "PASS"),
            ddl_align
        ]) else "FAIL"
        remarks="; ".join(remark)

        # AI queue (optional, non-blocking)
        if ai_enabled:
            trans_text = vals[ix_trans] if ix_trans is not None and ix_trans < len(vals) else ""
            if s(trans_text)=="" or len(trans_text.split())<3:
                ai_candidates.append({
                    "row": r_idx,
                    "type": "Transformation_Quality",
                    "dest_table": dest_table,
                    "dest_col": dest_col,
                    "text": trans_text
                })
            reason = vals[ix_reason] if ix_reason is not None and ix_reason < len(vals) else ""
            if mig=="NO" and (s(reason)=="" or "TBD" in up(reason) or "CHECK" in up(reason)):
                ai_candidates.append({
                    "row": r_idx,
                    "type": "Reason_Clarity",
                    "dest_table": dest_table,
                    "dest_col": dest_col,
                    "text": reason
                })

        out_ws.append(vals + [
            mapping_check, change_check, table_consistency,
            ddl_align, ref_check, ver_check,
            status, remarks, change_log
        ])

    # Missing_Tables_Report
    if table_details:
        missing = sorted(t for t in table_details if t not in seen_dest_tables)
        miss_ws = out_wb.create_sheet("Missing_Tables_Report")
        miss_ws.append(["Table Name","Comment"])
        if missing:
            for t in missing:
                miss_ws.append([t, "Present in Table Details but not found in Baseline Data Model"])
            print(f"[INFO] Missing_Tables_Report: {len(missing)} missing tables listed.")
        else:
            miss_ws.append(["(none)","All tables in Table Details are present in Baseline Data Model"])

    # Summary_Stats
    sum_ws = out_wb.create_sheet("Summary_Stats")
    sum_ws.append(["Rule","PASS","FAIL","SKIPPED","Total"])
    for r,counts in list({
        "Mapping_Check": stats["Mapping_Check"],
        "ChangeTracking_Check": stats["ChangeTracking_Check"],
        "TableConsistency_Check": stats["TableConsistency_Check"],
        "DDLAlignment_Check": stats["DDLAlignment_Check"],
        "ReferenceSubset_Check": stats["ReferenceSubset_Check"],
        "VersionDiff_Check": stats["VersionDiff_Check"]
    }.items()):
        p=counts.get("PASS",0); f=counts.get("FAIL",0); sk=counts.get("SKIPPED",0)
        sum_ws.append([r, p, f, sk, p+f+sk])

    # Save workbook
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(args.out)

    # Save AI queue (if any)
    if ai_enabled:
        out_dir = Path(args.out).parent
        ai_file = out_dir / "ai_pending.json"
        if ai_candidates:
            ai_file.write_text(json.dumps(ai_candidates, indent=2), encoding="utf-8")
            print(f"[AI] {len(ai_candidates)} candidates saved to {ai_file}")
        else:
            print("[AI] No AI candidates detected.")

    print(f"[OK] Output saved to {args.out}")

if __name__=="__main__":
    main()
