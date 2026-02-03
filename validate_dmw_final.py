#!/usr/bin/env python3
import argparse, traceback, logging, re
from pathlib import Path
from typing import Dict, List, Tuple, Set, Optional, Iterable

from openpyxl import load_workbook, Workbook
from cfg import PATHS

# ----------------------------------------------------
# Logging
# ----------------------------------------------------
raw_log = PATHS.get("logs", "./logs")
LOG_DIR = Path("./logs") if str(raw_log).startswith("/app") else Path(raw_log)
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_PATH = LOG_DIR / "dmw_validator.log"

logging.basicConfig(
    filename=str(LOG_PATH),
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# ----------------------------------------------------
# Helpers
# ----------------------------------------------------
def s(v):
    return "" if v is None else str(v).strip()

def norm_col(name: str) -> str:
    if not name:
        return ""
    n = str(name).upper().strip().replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", n)

def is_na(v: str) -> bool:
    t = s(v).upper()
    return t in ("", "NA", "N/A", "NIL")

def yn(v: str) -> str:
    t = s(v).upper()
    if t in ("Y", "YES", "TRUE", "1"):
        return "YES"
    if t in ("N", "NO", "FALSE", "0"):
        return "NO"
    return ""

def detect_header_row_flexible(ws, *, min_non_empty: int = 10, max_scan: int = 30, default_row: int = 2) -> int:
    """Heuristic: find first row with >= min_non_empty non-empty cells."""
    try:
        for r in range(1, max_scan + 1):
            row = next(ws.iter_rows(min_row=r, max_row=r))
            non_empty = sum(1 for c in row if c.value not in (None, ""))
            if non_empty >= min_non_empty:
                return r
    except Exception:
        pass
    return default_row

def build_header_index(ws, header_row: int):
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    columns = [s(v) for v in header_cells]
    while columns and columns[-1] == "":
        columns.pop()
    # NOTE: duplicates exist in real DMW (e.g., "Data Type" appears twice). We keep lists of indices.
    lookup: Dict[str, List[int]] = {}
    for i, c in enumerate(columns):
        k = norm_col(c)
        if not k:
            continue
        lookup.setdefault(k, []).append(i)
    return columns, lookup

# Canonical -> aliases seen in real IRAS/GDS style DMWs
HEADER_ALIASES: Dict[str, List[str]] = {
    # Rule1
    "MIGRATING COLUMN": [
        "MIGRATING OR NOT (YES/NO)",
        "MIGRATING? (YES/NO)",
        "MIGRATING (YES/NO)",
    ],
    "DESTINATION DATA TYPE": [
        "DATA TYPE",  # appears twice (source and destination) -> we choose destination via anchor
        "DEST DATA TYPE",
    ],
    "DESTINATION DATA LENGTH": [
        "DESTINATION LENGTH",
        "MAX LENGTH",  # appears twice -> choose destination via anchor
        "LENGTH",
    ],
    "DESTINATION NULLABLE": [
        "IS IT NULLABLE? YES/NO",
        "NULLABLE? (YES/NO)",
        "NULLABLE",
        "IS NULLABLE",
    ],
    "TRANSFORMATION LOGIC": [
        "TRANSFORMATION DESCRIPTION",
        "TRANSFORMATION",
        "TRANSFORMATION RULE",
    ],
    # Rule2
    "INTRODUCED SPRINT": [
        "INTRODUCED SPRINT (FOR DATA MIGRATION SPRINT)",
        "INTRODUCED SPRINT/PASS",
    ],
    "LAST UPDATED SPRINT": [
        "LAST UPDATED IN SPRINT/PASS",
        "LAST UPDATED SPRINT/PASS",
        "LAST UPDATED SPRINT",
    ],
    "CHANGE LOG": [
        "CHANG LOG (FOR DATA MIGRATION REFERENCE)",  # typo in v1_3
        "CHANGE LOG (FOR DATA MIGRATION REFERENCE)",
        "CHANGELOG",
    ],
    # Core columns (keep a few safe fallbacks)
    "SOURCE COLUMN NAME": ["SOURCE COLUMN"],
    "DESTINATION COLUMN NAME": ["DEST COLUMN NAME", "DESTINATION COLUMN"],
    "DESTINATION TABLE": ["DEST TABLE", "DESTINATION TABLE NAME"],
    "SOURCE TABLE": ["SOURCE TABLE NAME"],
}

def _collect_candidate_indices(lookup: Dict[str, List[int]], canonical: str) -> List[int]:
    names: List[str] = [canonical] + HEADER_ALIASES.get(norm_col(canonical), [])
    idxs: List[int] = []
    seen = set()
    for nm in names:
        for i in lookup.get(norm_col(nm), []):
            if i not in seen:
                idxs.append(i)
                seen.add(i)
    return sorted(idxs)

def resolve_col(lookup: Dict[str, List[int]],
                canonical: str,
                *,
                prefer_after: Optional[int] = None,
                prefer_before: Optional[int] = None) -> Optional[int]:
    """
    Resolve a column index by canonical name + aliases.
    Handles duplicates by choosing the closest match after/before an anchor when provided.
    """
    candidates = _collect_candidate_indices(lookup, canonical)
    if not candidates:
        return None

    if prefer_after is not None:
        after = [i for i in candidates if i > prefer_after]
        if after:
            return min(after)

    if prefer_before is not None:
        before = [i for i in candidates if i < prefer_before]
        if before:
            return max(before)

    return candidates[0]

def any_strikethrough(row_cells) -> bool:
    """Best-effort strikethrough detection."""
    try:
        for c in row_cells:
            f = getattr(c, "font", None)
            if f is not None and getattr(f, "strike", False):
                return True
    except Exception:
        return False
    return False

def normalize_sql_type(t: str) -> Tuple[str, str]:
    tt = s(t).upper()
    m = re.match(r"^([A-Z0-9_]+)\s*(\([^)]*\))?\s*$", tt)
    if not m:
        return (tt, "")
    base = m.group(1) or tt
    params = (m.group(2) or "").replace(" ", "")
    return (base, params)

def type_compatible(dmw_type: str, ddl_type: str) -> bool:
    if is_na(dmw_type) or is_na(ddl_type):
        return False
    b1, p1 = normalize_sql_type(dmw_type)
    b2, p2 = normalize_sql_type(ddl_type)
    if b1 != b2:
        return False
    if p1 and p2:
        return p1 == p2
    return True

def normalize_nullable(v: str) -> str:
    t = s(v).upper()
    if t in ("NOT NULL", "NO", "N", "FALSE", "0", "NN"):
        return "NOT NULL"
    if t in ("NULL", "YES", "Y", "TRUE", "1"):
        return "NULL"
    return ""

# ----------------------------------------------------
# DDL PARSER (SQL Server / Azure SQL)
# ----------------------------------------------------
def sniff_encoding(raw: bytes) -> str:
    if raw.startswith(b"\xff\xfe"):
        return "utf-16-le"
    if raw.startswith(b"\xfe\xff"):
        return "utf-16-be"
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "latin-1"

def parse_ddl(path: str) -> Dict[str, Dict[str, Dict[str, str]]]:
    """
    Returns:
      tables[table_upper][col_upper] = {"type": "...", "nullable": "NULL|NOT NULL|"}
    """
    raw = Path(path).read_bytes()
    enc = sniff_encoding(raw)
    ddl_text = raw.decode(enc, errors="ignore")

    tables: Dict[str, Dict[str, Dict[str, str]]] = {}

    create_re = re.compile(
        r"CREATE\s+TABLE\s+(?:\[(?P<schema>\w+)\]\.)?\[(?P<table>\w+)\]\s*\(",
        re.IGNORECASE
    )

    col_re = re.compile(
        r"^\s*\[(?P<col>\w+)\]\s+\[?(?P<type>[A-Za-z0-9_]+(?:\s*\([^)]*\))?)\]?(?P<rest>.*)$",
        re.IGNORECASE
    )

    for m in create_re.finditer(ddl_text):
        tbl = s(m.group("table")).upper()
        start = m.end()

        depth = 1
        i = start
        while i < len(ddl_text) and depth > 0:
            ch = ddl_text[i]
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            i += 1

        block = ddl_text[start:i - 1]
        cols: Dict[str, Dict[str, str]] = {}

        for line in block.splitlines():
            l = line.strip()
            if not l:
                continue
            up = l.upper()

            if up.startswith((
                "CONSTRAINT", "PRIMARY KEY", "FOREIGN KEY", "CHECK",
                "UNIQUE", "INDEX", "PERIOD FOR SYSTEM_TIME", "WITH"
            )):
                continue

            m2 = col_re.match(l)
            if not m2:
                continue
            # ❌ skip constraints & non-columns explicitly
            if up.startswith((
                "CONSTRAINT",
                "PRIMARY KEY",
                "FOREIGN KEY",
                "CHECK",
                "UNIQUE",
                "INDEX",
                "PERIOD FOR SYSTEM_TIME",
                "WITH"
            )):
                continue
            c = s(m2.group("col")).upper()
            t = s(m2.group("type")).upper()
            rest = s(m2.group("rest")).upper()
          
            if re.search(r"\bAS\s*\(", up):
                continue

            rest = re.sub(r"\bCOLLATE\b.*", "", rest)
            rest = re.sub(r"\bIDENTITY\s*\(.*?\)", "", rest)

            nullable = ""
            if "NOT NULL" in rest:
                nullable = "NOT NULL"
            elif re.search(r"\bNULL\b", rest):
                nullable = "NULL"

            cols[c] = {"type": t, "nullable": nullable}

        tables[tbl] = cols

    return tables

def ddl_diff(prev: Dict[str, Dict[str, Dict[str, str]]],
             curr: Dict[str, Dict[str, Dict[str, str]]]):
    added_tables = sorted(set(curr.keys()) - set(prev.keys()))
    removed_tables = sorted(set(prev.keys()) - set(curr.keys()))
    common_tables = sorted(set(curr.keys()) & set(prev.keys()))

    added_cols: List[Tuple[str, str, str, str]] = []
    removed_cols: List[Tuple[str, str, str, str]] = []
    changed_cols: List[Tuple[str, str, str, str, str, str]] = []

    for t in common_tables:
        pcols = prev[t]
        ccols = curr[t]
        pset = set(pcols.keys())
        cset = set(ccols.keys())

        for c in sorted(cset - pset):
            added_cols.append((t, c, ccols[c].get("type", ""), ccols[c].get("nullable", "")))
        for c in sorted(pset - cset):
            removed_cols.append((t, c, pcols[c].get("type", ""), pcols[c].get("nullable", "")))
        for c in sorted(pset & cset):
            pt = pcols[c].get("type", "")
            ct = ccols[c].get("type", "")
            pn = pcols[c].get("nullable", "")
            cn = ccols[c].get("nullable", "")
            if normalize_sql_type(pt) != normalize_sql_type(ct) or (pn or "") != (cn or ""):
                changed_cols.append((t, c, pt, pn, ct, cn))

    return added_tables, removed_tables, added_cols, removed_cols, changed_cols

# ----------------------------------------------------
# Rule1 / Rule2 (row-level)
# ----------------------------------------------------
def rule1_check(vals: List[str], *,
                mig_i: Optional[int],
                rsn_i: Optional[int],
                dtype_i: Optional[int],
                dlen_i: Optional[int],
                dnull_i: Optional[int],
                trans_i: Optional[int]) -> Tuple[str, str]:
    migrating = yn(vals[mig_i] if mig_i is not None and mig_i < len(vals) else "")
    reason    = vals[rsn_i] if rsn_i is not None and rsn_i < len(vals) else ""
    dtype     = vals[dtype_i] if dtype_i is not None and dtype_i < len(vals) else ""
    dlen      = vals[dlen_i]  if dlen_i  is not None and dlen_i  < len(vals) else ""
    dnull     = vals[dnull_i] if dnull_i is not None and dnull_i < len(vals) else ""
    trans     = vals[trans_i] if trans_i is not None and trans_i < len(vals) else ""

    if migrating == "NO":
        if rsn_i is not None and is_na(reason):
            return ("FAIL", "Rule1: Reason for not migrating is mandatory when Migrating Column = No")
        return ("PASS", "")

    if migrating == "YES":
        missing = []
        if dtype_i is not None and is_na(dtype):
            missing.append("datatype")
        if dnull_i is not None and is_na(dnull):
            missing.append("nullable")
        if trans_i is not None and is_na(trans):
            missing.append("transformation logic")

        # Length enforcement only when needed and dtype lacks params
        if dlen_i is not None and not is_na(dtype):
            base, params = normalize_sql_type(dtype)
            needs_len = base in ("CHAR", "NCHAR", "VARCHAR", "NVARCHAR", "BINARY", "VARBINARY", "DECIMAL", "NUMERIC")
            if needs_len and not params and is_na(dlen):
                missing.append("length")

        if missing:
            return ("FAIL", f"Rule1: Missing destination fields: {', '.join(missing)}")
        return ("PASS", "")

    # If migrating flag isn't recognized, do not fail by default.
    return ("PASS", "")

def rule2_check(vals: List[str], *,
                intro_i: Optional[int],
                last_i: Optional[int],
                log_i: Optional[int]) -> Tuple[str, str]:
    if intro_i is None or last_i is None or log_i is None:
        return ("PASS", "")

    intro = vals[intro_i] if intro_i < len(vals) else ""
    last  = vals[last_i]  if last_i  < len(vals) else ""
    clog  = vals[log_i]   if log_i   < len(vals) else ""

    if not is_na(intro) and not is_na(last) and s(intro).upper() != s(last).upper():
        if is_na(clog):
            return ("FAIL", "Rule2: Change log required when Introduced Sprint differs from Last Updated Sprint")
    return ("PASS", "")

# ----------------------------------------------------
# Rule5/6 helpers (need same alias/dup-safe column resolution)
# ----------------------------------------------------
def _load_sheet_header(ws) -> Tuple[int, List[str], Dict[str, List[int]]]:
    hr = detect_header_row_flexible(ws, min_non_empty=1, max_scan=30, default_row=1)
    cols, lookup = build_header_index(ws, hr)
    return hr, cols, lookup

def load_dmw_dest_keys(path: str) -> Dict[str, Set[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hr, cols, lookup = _load_sheet_header(ws)
    start = hr + 1

    dt_i = resolve_col(lookup, "Destination Table")
    dc_i = resolve_col(lookup, "Destination Column Name", prefer_after=dt_i if dt_i is not None else None)

    out: Dict[str, Set[str]] = {}
    if dt_i is None or dc_i is None:
        wb.close()
        return out

    for r in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True):
        if r is None:
            break
        vals = [s(v) for v in r[:len(cols)]]
        if all(v == "" for v in vals):
            break
        DT = vals[dt_i] if dt_i < len(vals) else ""
        DC = vals[dc_i] if dc_i < len(vals) else ""
        if is_na(DT) or is_na(DC):
            continue
        out.setdefault(s(DT).upper(), set()).add(s(DC).upper())

    wb.close()
    return out

def load_dmw_dest_defs(path: str) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    For Rule6B (attribute drift, INFO only).
    Returns:
      defs[(T,C)] = {"type": "...", "nullable": "...", "transform": "..."}
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hr, cols, lookup = _load_sheet_header(ws)
    start = hr + 1

    dt_i = resolve_col(lookup, "Destination Table")
    dc_i = resolve_col(lookup, "Destination Column Name", prefer_after=dt_i if dt_i is not None else None)

    # Destination-specific fields can be duplicated ("Data Type", "Max Length") -> anchor after dest column
    dtype_i = resolve_col(lookup, "Destination Data Type", prefer_after=dc_i)
    dlen_i  = resolve_col(lookup, "Destination Data Length", prefer_after=dc_i)
    dnull_i = resolve_col(lookup, "Destination Nullable", prefer_after=dc_i)
    trans_i = resolve_col(lookup, "Transformation Logic", prefer_after=dc_i)

    out: Dict[Tuple[str, str], Dict[str, str]] = {}
    if dt_i is None or dc_i is None:
        wb.close()
        return out

    for r in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True):
        if r is None:
            break
        vals = [s(v) for v in r[:len(cols)]]
        if all(v == "" for v in vals):
            break

        DT = vals[dt_i] if dt_i < len(vals) else ""
        DC = vals[dc_i] if dc_i < len(vals) else ""
        if is_na(DT) or is_na(DC):
            continue

        dmw_type = vals[dtype_i] if dtype_i is not None and dtype_i < len(vals) else ""
        dmw_len  = vals[dlen_i]  if dlen_i  is not None and dlen_i  < len(vals) else ""
        dmw_null = vals[dnull_i] if dnull_i is not None and dnull_i < len(vals) else ""
        dmw_tran = vals[trans_i] if trans_i is not None and trans_i < len(vals) else ""

        dmw_type_full = s(dmw_type).upper()
        if dmw_len and "(" not in dmw_type_full and ")" not in dmw_type_full:
            dmw_type_full = f"{dmw_type_full}({s(dmw_len)})"

        out[(s(DT).upper(), s(DC).upper())] = {
            "type": dmw_type_full,
            "nullable": normalize_nullable(dmw_null),
            "transform": s(dmw_tran),
        }

    wb.close()
    return out

def dmw_drift(prev: Dict[str, Set[str]], curr: Dict[str, Set[str]]):
    prev_keys = {(t, c) for t, cols in prev.items() for c in cols}
    curr_keys = {(t, c) for t, cols in curr.items() for c in cols}
    added = sorted(curr_keys - prev_keys)
    removed = sorted(prev_keys - curr_keys)
    modified: List[Tuple[str, str, str]] = []
    return added, removed, modified

# ----------------------------------------------------
# MAIN VALIDATION
# ----------------------------------------------------
def validate(dmw_xlsx, ddl_sql, out_xlsx, ai_cfg, prev_dmw=None, prev_ddl=None, ref_dmw=None, master_dmw=None):
    #ddl_curr = parse_ddl(ddl_sql)
    from parse_ddl_v2 import parse_ddl_v2
    ddl_curr = parse_ddl_v2(ddl_sql)
    ddl_prev = parse_ddl(prev_ddl) if prev_ddl else None

    prev_keys_by_table = load_dmw_dest_keys(prev_dmw) if prev_dmw else None
    prev_defs = load_dmw_dest_defs(prev_dmw) if prev_dmw else None

    master_keys = load_dmw_dest_keys(master_dmw) if master_dmw else None
    ref_keys = load_dmw_dest_keys(ref_dmw) if ref_dmw else None

    wb_data = load_workbook(dmw_xlsx, read_only=True, data_only=True)
    ws_data = wb_data.active

    header_row = detect_header_row_flexible(ws_data, min_non_empty=10, max_scan=30, default_row=2)
    data_start = header_row + 1
    columns, lookup = build_header_index(ws_data, header_row)

    # Core identity
    st_i = resolve_col(lookup, "Source Table")
    sc_i = resolve_col(lookup, "Source Column Name", prefer_after=st_i if st_i is not None else None)
    dt_i = resolve_col(lookup, "Destination Table")
    dc_i = resolve_col(lookup, "Destination Column Name", prefer_after=dt_i if dt_i is not None else None)

    # Rule1 columns (destination fields are duplicated in real DMW; anchor after destination column)
    mig_i   = resolve_col(lookup, "Migrating Column", prefer_before=dt_i)
    rsn_i   = resolve_col(lookup, "Reason for Not Migrating", prefer_before=dt_i)
    dtype_i = resolve_col(lookup, "Destination Data Type", prefer_after=dc_i)
    dlen_i  = resolve_col(lookup, "Destination Data Length", prefer_after=dc_i)
    dnull_i = resolve_col(lookup, "Destination Nullable", prefer_after=dc_i)
    trans_i = resolve_col(lookup, "Transformation Logic", prefer_after=dc_i)

    # Rule2 columns
    intro_i = resolve_col(lookup, "Introduced Sprint")
    last_i  = resolve_col(lookup, "Last Updated Sprint")
    clog_i  = resolve_col(lookup, "Change Log")

    out_wb = Workbook()
    ws_main = out_wb.active
    ws_main.title = "Baseline Data Model_output"

    ws_r4 = out_wb.create_sheet("Rule4_DDL_Mismatch")
    ws_r4.append(["Table", "Column", "Issue", "Details"])

    ws_r3 = out_wb.create_sheet("Rule3_Table_Mismatch")
    ws_r3.append(["Table", "Issue", "Details"])

    ws_r5 = out_wb.create_sheet("Rule5_Ref_Master_Mismatch")
    ws_r5.append(["Master_Table", "Reference_Table", "Column", "Issue", "Details"])

    ws_r6 = out_wb.create_sheet("Rule6_DMW_Drift")
    ws_r6.append(["Dest_Table", "Dest_Column", "Issue", "Details"])

    ws_r7 = out_wb.create_sheet("Rule7_DDL_Drift")
    ws_r7.append(["Object", "Name", "Issue", "Details"])

    RULE_COLS = [
        "Rule1", "Rule2", "Rule3", "Rule4",
        "Rule5", "Rule6", "Rule7",
        "Validation_Status", "Validation_Remarks", "AI_Suggestion"
    ]
    ws_main.append(columns + RULE_COLS)

    dest_map: Dict[str, Set[str]] = {}
    dmw_defs: Dict[Tuple[str, str], Dict[str, str]] = {}    # current defs
    baseline_tables: Set[str] = set()
    curr_keys_by_table: Dict[str, Set[str]] = {}

    # -----------------------------
    # Single pass rows
    # -----------------------------
    for row_cells in ws_data.iter_rows(min_row=data_start, max_row=ws_data.max_row, values_only=False):
        vals = [s(c.value) for c in row_cells[:len(columns)]]
        if all(v == "" for v in vals):
            break

        # Strikethrough => N/A for all rules
        if any_strikethrough(row_cells):
            ws_main.append(vals + [
                "N/A", "N/A", "N/A", "N/A",
                "N/A", "N/A", "N/A",
                "N/A",
                "Strikethrough: field cancelled by Apps team",
                ""
            ])
            continue

        ST = vals[st_i] if st_i is not None and st_i < len(vals) else ""
        SC = vals[sc_i] if sc_i is not None and sc_i < len(vals) else ""
        DT = vals[dt_i] if dt_i is not None and dt_i < len(vals) else ""
        DC = vals[dc_i] if dc_i is not None and dc_i < len(vals) else ""

        source_na = is_na(ST) and is_na(SC)
        dest_na   = is_na(DT) and is_na(DC)

        # Helper row
        if source_na and dest_na:
            ws_main.append(vals + [
                "N/A", "N/A", "N/A", "N/A",
                "N/A", "N/A", "N/A",
                "N/A",
                "Source and Destination are NA — helper row",
                ""
            ])
            continue

        # Destination missing => Rule4 N/A (others run where applicable)
        if is_na(DT) or is_na(DC):
            r1, r1r = rule1_check(vals, mig_i=mig_i, rsn_i=rsn_i, dtype_i=dtype_i, dlen_i=dlen_i, dnull_i=dnull_i, trans_i=trans_i)
            r2, r2r = rule2_check(vals, intro_i=intro_i, last_i=last_i, log_i=clog_i)
            status = "FAIL" if (r1 == "FAIL" or r2 == "FAIL") else "N/A"
            remarks = "No Destination mapping — Rule4 skipped"
            extra = " | ".join([x for x in [r1r, r2r] if x])
            if extra:
                remarks = f"{remarks} | {extra}"

            ws_main.append(vals + [
                r1, r2, "N/A", "N/A",
                "N/A", "N/A", "N/A",
                status, remarks, ""
            ])
            continue

        # Normal row
        tblU = s(DT).upper()
        colU = s(DC).upper()

        baseline_tables.add(tblU)
        dest_map.setdefault(tblU, set()).add(colU)
        curr_keys_by_table.setdefault(tblU, set()).add(colU)

        r1, r1r = rule1_check(vals, mig_i=mig_i, rsn_i=rsn_i, dtype_i=dtype_i, dlen_i=dlen_i, dnull_i=dnull_i, trans_i=trans_i)
        r2, r2r = rule2_check(vals, intro_i=intro_i, last_i=last_i, log_i=clog_i)
        status = "FAIL" if (r1 == "FAIL" or r2 == "FAIL") else "PASS"
        remarks = " | ".join([x for x in [r1r, r2r] if x])

        # seed; Rule3/4/5/6/7 will be propagated later
        ws_main.append(vals + [
            r1, r2, "PASS", "PASS",
            "PASS", "PASS", "PASS",
            status, remarks, ""
        ])

        # store current DMW defs for Rule4A + Rule6B
        dmw_type = vals[dtype_i] if dtype_i is not None and dtype_i < len(vals) else ""
        dmw_len  = vals[dlen_i]  if dlen_i  is not None and dlen_i  < len(vals) else ""
        dmw_null = vals[dnull_i] if dnull_i is not None and dnull_i < len(vals) else ""
        dmw_tran = vals[trans_i] if trans_i is not None and trans_i < len(vals) else ""

        dmw_type_full = s(dmw_type).upper()
        if dmw_len and "(" not in dmw_type_full and ")" not in dmw_type_full:
            dmw_type_full = f"{dmw_type_full}({s(dmw_len)})"

        dmw_defs[(tblU, colU)] = {
            "type": dmw_type_full,
            "nullable": normalize_nullable(dmw_null),
            "transform": s(dmw_tran),
        }

    wb_data.close()
    # ------------------------------------------------
# Rule3: Baseline Data Model vs Table Details
# ------------------------------------------------
    # ------------------------------------------------
# Rule3: Baseline Data Model vs Table Details
# ------------------------------------------------
    try:
        wb_td = load_workbook(dmw_xlsx, read_only=True, data_only=True)
        ws_td = None

        # Locate "Table Details" sheet (robust match)
        for sn in wb_td.sheetnames:
            if norm_col(sn) == norm_col("TABLE DETAILS"):
                ws_td = wb_td[sn]
                break

        table_details_set: Set[str] = set()

        if ws_td is not None:
            hr = detect_header_row_flexible(
                ws_td,
                min_non_empty=1,
                max_scan=10,
                default_row=1
            )
            tcols, tlookup = build_header_index(ws_td, hr)
            start = hr + 1

            # Resolve table name column safely
            table_i = (
                resolve_col(tlookup, "Table Name")
                or resolve_col(tlookup, "Destination Table")
            )

            if table_i is None:
                for k, idxs in tlookup.items():
                    if "TABLE" in k:
                        table_i = idxs[0]
                        break

            if table_i is not None:
                for r in ws_td.iter_rows(
                    min_row=start,
                    max_row=ws_td.max_row,
                    values_only=True
                ):
                    if not r:
                        break

                    vals = [s(v) for v in r[:len(tcols)]]
                    if all(v == "" for v in vals):
                        break

                    tname = vals[table_i] if table_i < len(vals) else ""
                    if not is_na(tname):
                        table_details_set.add(s(tname).upper())

        # ------------------------------------------------
        # A️⃣ Baseline → Table Details missing
        # ------------------------------------------------
        rule3_missing_tables = sorted(baseline_tables - table_details_set)

        for t in rule3_missing_tables:
            ws_r3.append([
                t,
                "MISSING_IN_TABLE_DETAILS",
                "Destination table used in Baseline Data Model but not found in Table Details sheet"
            ])

        # ------------------------------------------------
        # B️⃣ Table Details → Baseline unused
        # ------------------------------------------------
        rule3_unused_tables = sorted(table_details_set - baseline_tables)

        for t in rule3_unused_tables:
            ws_r3.append([
                t,
                "UNUSED_IN_BASELINE",
                "Table listed in Table Details but not used in Baseline Data Model"
            ])

        wb_td.close()

    except Exception:
        logging.exception("Rule3 processing failed")


    # ------------------------------------------------
    # Rule4: DDL alignment (Rule4A + Rule4B)
    # ------------------------------------------------
 
    from parse_ddl_v2 import parse_ddl_v2

    ddl_curr = parse_ddl_v2(ddl_sql)

    mismatch_keys = set()
    table_has_rule4_issue = set()
    missing_in_dmw = []

    for tbl, ddl_cols in ddl_curr.items():
        tblU = tbl.upper()
        ddl_set = set(ddl_cols.keys())
        dmw_set = dest_map.get(tblU, set())

        # ------------------------------------------------
        # DMW_ONLY (exists in DMW, not in DDL)
        # ------------------------------------------------
        for col in sorted(dmw_set - ddl_set):
            ws_r4.append([
                tblU,
                col,
                "DMW_ONLY",
                "Destination column appears in DMW but not in DDL"
            ])
            mismatch_keys.add((tblU, col))
            table_has_rule4_issue.add(tblU)

        # ------------------------------------------------
        # Type / Nullable mismatches
        # ------------------------------------------------
        for col in sorted(dmw_set & ddl_set):
            ddl_def = ddl_cols.get(col, {})
            dmw_def = dmw_defs.get((tblU, col), {})

            ddl_type = ddl_def.get("type", "")
            dmw_type = dmw_def.get("type", "")

            ddl_null = ddl_def.get("nullable", "")
            dmw_null = dmw_def.get("nullable", "")

            if ddl_type and dmw_type and not type_compatible(dmw_type, ddl_type):
                ws_r4.append([
                    tblU,
                    col,
                    "TYPE_MISMATCH",
                    f"DMW type={dmw_type} vs DDL type={ddl_type}"
                ])
                mismatch_keys.add((tblU, col))
                table_has_rule4_issue.add(tblU)

            if ddl_null and dmw_null and ddl_null != dmw_null:
                ws_r4.append([
                    tblU,
                    col,
                    "NULLABLE_MISMATCH",
                    f"DMW nullable={dmw_null} vs DDL nullable={ddl_null}"
                ])
                mismatch_keys.add((tblU, col))
                table_has_rule4_issue.add(tblU)

        # ------------------------------------------------
        # MISSING_IN_DMW (exists in DDL, not in DMW)
        # ------------------------------------------------
        if dmw_set:
            for col in sorted(ddl_set - dmw_set):
                ws_r4.append([
                    tblU,
                    col,
                    "MISSING_IN_DMW",
                    "Column exists in DDL but not mapped in DMW"
                ])
                missing_in_dmw.append((tblU, col))
                table_has_rule4_issue.add(tblU)

    # ------------------------------------------------
    # Rule5: Reference tables subset of master tables
    # ------------------------------------------------
    rule5_fail_tables: Set[str] = set()
    if master_keys is not None and ref_keys is not None:
        for rt, rcols in ref_keys.items():
            mcols = master_keys.get(rt, set())
            if not mcols:
                for c in sorted(rcols):
                    ws_r5.append([rt, rt, c, "MASTER_TABLE_MISSING", "Reference table exists but master table not found"])
                rule5_fail_tables.add(rt)
                continue
            extra = sorted(rcols - mcols)
            for c in extra:
                ws_r5.append([rt, rt, c, "NOT_IN_MASTER", "Reference column not found in master table (ref must be subset)"])
                rule5_fail_tables.add(rt)

    # ------------------------------------------------
    # Rule6A: DMW drift (prev vs current) - FAIL on structural drift
    # Rule6B: Attribute drift (INFO only, do NOT fail baseline)
    # ------------------------------------------------
    drift_keys: Set[Tuple[str, str]] = set()
    if prev_keys_by_table is not None:
        added, removed, modified = dmw_drift(prev_keys_by_table, curr_keys_by_table)

        for (t, c) in added:
            ws_r6.append([t, c, "ADDED_IN_CURRENT", "Destination column exists in current DMW but not in previous frozen DMW"])
            drift_keys.add((t, c))

        for (t, c) in removed:
            ws_r6.append([t, c, "REMOVED_IN_CURRENT", "Destination column existed in previous frozen DMW but not in current DMW"])

        for (t, c, det) in modified:
            ws_r6.append([t, c, "MODIFIED", det])
            drift_keys.add((t, c))

        # Synthetic baseline rows for removed (so baseline shows FAIL evidence)
        if removed and dt_i is not None and dc_i is not None:
            for (t, c) in removed:
                data = [""] * len(columns)
                if st_i is not None and st_i < len(data):
                    data[st_i] = "NA"
                if sc_i is not None and sc_i < len(data):
                    data[sc_i] = "NA"
                if dt_i < len(data):
                    data[dt_i] = t
                if dc_i < len(data):
                    data[dc_i] = c
                ws_main.append(data + [
                    "N/A", "N/A", "PASS", "PASS",
                    "PASS", "FAIL", "PASS",
                    "FAIL",
                    "Rule6 drift – REMOVED_IN_CURRENT – see Rule6_DMW_Drift",
                    ""
                ])

        # Rule6B (INFO only): compare attributes for keys present in both prev and current
        if prev_defs is not None:
            for key, curr_def in dmw_defs.items():
                prev_def = prev_defs.get(key)
                if not prev_def:
                    continue
                t, c = key

                if normalize_sql_type(prev_def.get("type", "")) != normalize_sql_type(curr_def.get("type", "")):
                    ws_r6.append([t, c, "DATATYPE_CHANGED", f"Prev={prev_def.get('type','')} Curr={curr_def.get('type','')}"])

                if normalize_nullable(prev_def.get("nullable", "")) != normalize_nullable(curr_def.get("nullable", "")):
                    ws_r6.append([t, c, "NULLABLE_CHANGED", f"Prev={prev_def.get('nullable','')} Curr={curr_def.get('nullable','')}"])

                if s(prev_def.get("transform", "")) != s(curr_def.get("transform", "")):
                    ws_r6.append([t, c, "TRANSFORMATION_CHANGED", "Transformation logic changed"])

    # ------------------------------------------------
    # Rule7: DDL drift (prev vs current)
    # ------------------------------------------------
    if ddl_prev is not None:
        added_tables, removed_tables, added_cols, removed_cols, changed_cols = ddl_diff(ddl_prev, ddl_curr)

        for t in added_tables:
            ws_r7.append(["TABLE", t, "ADDED_IN_CURRENT", "Table exists in current DDL but not in previous DDL"])
        for t in removed_tables:
            ws_r7.append(["TABLE", t, "REMOVED_IN_CURRENT", "Table exists in previous DDL but not in current DDL"])
        for (t, c, typ, nul) in added_cols:
            ws_r7.append(["COLUMN", f"{t}.{c}", "ADDED_IN_CURRENT", f"type={typ} nullable={nul}"])
        for (t, c, typ, nul) in removed_cols:
            ws_r7.append(["COLUMN", f"{t}.{c}", "REMOVED_IN_CURRENT", f"type={typ} nullable={nul}"])
        for (t, c, pt, pn, ct, cn) in changed_cols:
            ws_r7.append(["COLUMN", f"{t}.{c}", "MODIFIED", f"prev type={pt} nullable={pn} | curr type={ct} nullable={cn}"])

    # ------------------------------------------------
    # Propagate Rule3/4/5/6/7 to baseline (single rewrite)
    # ------------------------------------------------
    rule_cols_n = len(RULE_COLS)
    all_rows = list(ws_main.iter_rows(min_row=2, values_only=True))
    ws_main.delete_rows(2, ws_main.max_row)

    rule3_has_issues = ws_r3.max_row > 1
    rule7_has_issues = ws_r7.max_row > 1  # sheet-only

    for r in all_rows:
        if r is None or len(r) < rule_cols_n:
            continue

        data = list(r[:-rule_cols_n])
        r1, r2, r3, r4, r5, r6, r7, status, remarks, ai = r[-rule_cols_n:]

        DT = s(data[dt_i]) if dt_i is not None and dt_i < len(data) else ""
        DC = s(data[dc_i]) if dc_i is not None and dc_i < len(data) else ""
        tblU = DT.upper()
        key = (tblU, DC.upper())

        # Rule3: fail rows that have a destination table if table details mismatch exists
       # if rule3_has_issues and not is_na(DT):
        #    r3 = "INFO"
         #   remarks = (remarks + " | " if remarks else "") + "Rule3 table mismatch – see Rule3_Table_Mismatch"

        rule3_fail_tables = {row[0] for row in ws_r3.iter_rows(min_row=2, values_only=True)}

        # later, per row
        if tblU in rule3_fail_tables:
            r3 = "FAIL"
            status = "FAIL"
            remarks = (remarks + " | " if remarks else "") + \
                    "Rule3 mismatch – see Rule3_Table_Mismatch"
        else:
            r3 = "PASS"


                # Rule4: exact mismatch OR table-level escalation if table has any Rule4 issues
        if r4 != "N/A":
            if key in mismatch_keys or tblU in table_has_rule4_issue:
                r4 = "FAIL"
                status = "FAIL"
                remarks = (remarks + " | " if remarks else "") + "Rule4 mismatch – see Rule4_DDL_Mismatch"
            else:
                r4 = "PASS"

        # Rule5: if table flagged in rule5_fail_tables, fail rows for that table
        if master_keys is not None and ref_keys is not None:
            if tblU in rule5_fail_tables:
                r5 = "FAIL"
                status = "FAIL"
                remarks = (remarks + " | " if remarks else "") + "Rule5 mismatch – see Rule5_Ref_Master_Mismatch"
            else:
                r5 = "PASS"

        # Rule6A only: structural drift causes FAIL; Rule6B does NOT fail baseline
        if prev_keys_by_table is not None:
            if key in drift_keys:
                r6 = "FAIL"
                status = "FAIL"
                remarks = (remarks + " | " if remarks else "") + "Rule6 drift – see Rule6_DMW_Drift"
            else:
                r6 = "PASS"

        # Rule7: sheet-only
        if ddl_prev is not None:
            r7 = "PASS" if not rule7_has_issues else "PASS"

        ws_main.append(data + [r1, r2, r3, r4, r5, r6, r7, status, remarks, ai])

    out_wb.save(out_xlsx)
    print(f"[OK] Validation completed → {out_xlsx}")
    logging.info(f"Validation completed → {out_xlsx}")

# ----------------------------------------------------
# CLI
# ----------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dmw-xlsx", required=True)
    ap.add_argument("--ddl-sql", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--enable-ai", action="store_true")
    ap.add_argument("--prev-dmw", default=None)
    ap.add_argument("--prev-ddl", default=None)
    ap.add_argument("--ref-dmw", default=None)
    ap.add_argument("--master-dmw", default=None)

    args = ap.parse_args()
    ai_cfg = {"enabled": args.enable_ai}

    try:
        validate(
            args.dmw_xlsx,
            args.ddl_sql,
            args.out,
            ai_cfg,
            prev_dmw=args.prev_dmw,
            prev_ddl=args.prev_ddl,
            ref_dmw=args.ref_dmw,
            master_dmw=args.master_dmw
        )
    except Exception:
        traceback.print_exc()
        logging.exception("FATAL ERROR")

if __name__ == "__main__":
    main()