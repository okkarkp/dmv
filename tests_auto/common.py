#!/usr/bin/env python3
import os, shutil, subprocess, tempfile
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

VALIDATOR = Path("/opt/oss-migrate/llm-planner/validate_dmw_final.py")

# ----------------------------
# Minimal DMW generator
# ----------------------------
DMW_HEADERS = [
    "Source Table", "Source Column Name",
    "Destination Table", "Destination Column Name",

    # Rule1 fields
    "Migrating Column", "Reason for Not Migrating",
    "Destination Data Type", "Destination Data Length",
    "Destination Nullable", "Transformation Logic",

    # Rule2 fields
    "Introduced Sprint", "Last Updated Sprint", "Change Log",
]

def make_dmw_xlsx(path: Path, rows: List[Dict[str, Any]], *,
                 add_table_details: Optional[List[str]] = None,
                 strike_row_indexes: Optional[List[int]] = None) -> None:
    """
    rows: list of dict keyed by DMW_HEADERS.
    add_table_details: list of table names to add into "Table Details" sheet.
    strike_row_indexes: 0-based indexes into 'rows' to apply strikethrough to ALL cells in that row.
      NOTE: Your validator loads DMW in read_only mode; strikethrough detection is best-effort.
            This test may be skipped if strike cannot be detected in your openpyxl/runtime.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Baseline Data Model"
    ws.append(DMW_HEADERS)

    strike_row_indexes = set(strike_row_indexes or [])

    for i, r in enumerate(rows):
        values = [str(r.get(h, "") or "") for h in DMW_HEADERS]
        ws.append(values)
        if i in strike_row_indexes:
            # Apply strike to every cell in the appended row (row index: +2 because header)
            excel_row = i + 2
            for col in range(1, len(DMW_HEADERS) + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.font = Font(strike=True)

    if add_table_details is not None:
        ws2 = wb.create_sheet("Table Details")
        # A very simple table details layout
        ws2.append(["Table Name", "Remark"])
        for t in add_table_details:
            ws2.append([t, ""])

    wb.save(path)

# ----------------------------
# Minimal DDL generator
# ----------------------------
def make_ddl_sql(path: Path, tables: Dict[str, Dict[str, str]]) -> None:
    """
    tables = { "T1": {"C1": "INT NOT NULL", "C2": "NVARCHAR(50) NULL"} }
    Generates Azure SQL-ish CREATE TABLE blocks.
    """
    lines: List[str] = []
    for t, cols in tables.items():
        lines.append(f"CREATE TABLE [{t}] (")
        col_lines = []
        for c, spec in cols.items():
            col_lines.append(f"  [{c}] {spec}")
        # trailing commas are common; keep safe by joining with ",\n"
        lines.append(",\n".join(col_lines))
        lines.append(");")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")

# ----------------------------
# Run validator
# ----------------------------
def run_validator(*, dmw: Path, ddl: Path, out: Path,
                  prev_dmw: Optional[Path] = None,
                  prev_ddl: Optional[Path] = None,
                  ref_dmw: Optional[Path] = None,
                  master_dmw: Optional[Path] = None) -> None:
    cmd = ["python3", str(VALIDATOR), "--dmw-xlsx", str(dmw), "--ddl-sql", str(ddl), "--out", str(out)]
    if prev_dmw:
        cmd += ["--prev-dmw", str(prev_dmw)]
    if prev_ddl:
        cmd += ["--prev-ddl", str(prev_ddl)]
    if ref_dmw:
        cmd += ["--ref-dmw", str(ref_dmw)]
    if master_dmw:
        cmd += ["--master-dmw", str(master_dmw)]

    subprocess.check_call(cmd)

# ----------------------------
# Output readers + assertions
# ----------------------------
def read_sheet_rows(xlsx: Path, sheet: str) -> List[Tuple[Any, ...]]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise AssertionError(f"Missing sheet: {sheet}. Present: {wb.sheetnames}")
    ws = wb[sheet]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows

def find_col_index(header_row: Tuple[Any, ...], col_name: str) -> int:
    target = str(col_name).strip()
    for i, v in enumerate(header_row):
        if str(v or "").strip() == target:
            return i
    raise AssertionError(f"Column '{col_name}' not found in header: {header_row}")

def assert_any_row_has_value(rows: List[Tuple[Any, ...]], col_name: str, expected: str) -> None:
    header = rows[0]
    idx = find_col_index(header, col_name)
    for r in rows[1:]:
        if str(r[idx] or "").strip() == expected:
            return
    raise AssertionError(f"Expected value '{expected}' not found in column '{col_name}'")

def assert_any_row_matches(rows: List[Tuple[Any, ...]], predicate, msg: str) -> None:
    for r in rows[1:]:
        if predicate(r):
            return
    raise AssertionError(msg)

def assert_sheet_has_issue(xlsx: Path, sheet: str, issue_col_name: str, expected_issue: str) -> None:
    rows = read_sheet_rows(xlsx, sheet)
    header = rows[0]
    idx = find_col_index(header, issue_col_name)
    for r in rows[1:]:
        if str(r[idx] or "").strip() == expected_issue:
            return
    raise AssertionError(f"Expected issue '{expected_issue}' not found in {sheet}.{issue_col_name}")

# ----------------------------
# Test workspace helper
# ----------------------------
class Workdir:
    def __init__(self, prefix: str = "dmwtest_"):
        self.root = Path(tempfile.mkdtemp(prefix=prefix))
    def p(self, name: str) -> Path:
        return self.root / name
    def cleanup(self):
        shutil.rmtree(self.root, ignore_errors=True)
