from openpyxl import load_workbook
wb = load_workbook("/app/outputs/Withholding_Validation_AI.xlsx", read_only=True, data_only=True)
ws = wb["Baseline Data Model_output"]
headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
col_status = [i for i,h in enumerate(headers) if "Status" in str(h)]
print("Columns:", headers[col_status[0]] if col_status else "No Status column found")

counts = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or not any(row): continue
    status = str(row[col_status[0]]) if col_status else ""
    counts[status] = counts.get(status, 0) + 1
print("Validation Status counts:", counts)
