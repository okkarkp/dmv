#!/usr/bin/env python3
import argparse, traceback, re
from pathlib import Path
from openpyxl import load_workbook
from llama_cpp import Llama

def s(v): return "" if v is None else str(v).strip()
def up(v): return s(v).upper()

def ai_comment(prompt, ai_cfg):
    if not ai_cfg.get("enabled"): return ""
    try:
        from requests import post
        payload = {"prompt": prompt, "max_tokens": 60}
        url = "http://127.0.0.1:8080/v1/completions"
        j = post(url, json=payload, timeout=20).json()
        if "choices" in j: return "-- AI: " + j["choices"][0]["text"].strip()
    except Exception as e:
        return f"-- [AI error: {e}]"
    return ""

def build_ddl_index(sql_text: str):
    ddl={}
    for m in re.finditer(r"CREATE\\s+TABLE", sql_text, re.I):
        end=sql_text.find(";",m.start())
        end=len(sql_text) if end==-1 else end+1
        stmt=sql_text[m.start():end]
        tm=re.search(r"CREATE\\s+TABLE\\s+([^\\(\\s]+)",stmt,re.I)
        if not tm: continue
        tname=up(tm.group(1).split(".")[-1].strip('[]"`'))
        ddl[tname]=stmt
    return ddl

def generate(dmw_xlsx:Path, ddl_sql:Path, out_dir:Path, ai_cfg:dict):
    print(f"[INFO] Generating DDL from {dmw_xlsx.name}")
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(dmw_xlsx, data_only=True)
    ws = wb.active
    headers = [s(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx_tbl=headers.index("Destination Table") if "Destination Table" in headers else None
    idx_col=headers.index("Destination Column Name") if "Destination Column Name" in headers else None
    idx_dtype=headers.index("Data Type") if "Data Type" in headers else None
    idx_len=headers.index("Max Length") if "Max Length" in headers else None
    idx_null=headers.index("Is it Nullable? Yes/No") if "Is it Nullable? Yes/No" in headers else None

    ddl_idx=build_ddl_index(ddl_sql.read_text(encoding="utf-8",errors="ignore"))
    tables={}
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not any(row): continue
        dt=up(row[idx_tbl]) if idx_tbl is not None else ""
        dc=up(row[idx_col]) if idx_col is not None else ""
        dtype=s(row[idx_dtype]); length=s(row[idx_len]); null=s(row[idx_null])
        if not dt or not dc: continue
        ddl_type=f"{dtype}({length})" if length else dtype
        nullable="NULL" if null.upper().startswith("Y") else "NOT NULL"
        tables.setdefault(dt,[]).append((dc,ddl_type,nullable))

    for tname, cols in tables.items():
        fpath=out_dir/f"DDL_{tname}.sql"
        with fpath.open("w") as f:
            if tname in ddl_idx:
                f.write(f"-- Updating {tname}\\n")
                for dc,ddl_type,nullable in cols:
                    f.write(f"ALTER TABLE {tname} ADD COLUMN {dc} {ddl_type} {nullable};\\n")
                    if ai_cfg.get("enabled"):
                        prompt=f"Suggest improvement for column {dc} in table {tname}."
                        f.write(ai_comment(prompt,ai_cfg)+"\\n")
            else:
                f.write(f"CREATE TABLE {tname} (\\n")
                for i,(dc,ddl_type,nullable) in enumerate(cols):
                    comma="," if i<len(cols)-1 else ""
                    f.write(f"  {dc} {ddl_type} {nullable}{comma}\\n")
                f.write(");\\n")
                if ai_cfg.get("enabled"):
                    prompt=f"Suggest indexes or constraints for table {tname}."
                    f.write(ai_comment(prompt,ai_cfg)+"\\n")
        print(f"[OK] {fpath.name}")
    print(f"[DONE] Generated all DDLs → {out_dir}")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--dmw-xlsx",required=True)
    ap.add_argument("--ddl-sql",required=True)
    ap.add_argument("--out-dir",required=True)
    ap.add_argument("--enable-ai-sql",action="store_true")
    args=ap.parse_args()
    ai_cfg={"enabled":args.enable_ai_sql}
    try: generate(Path(args.dmw_xlsx),Path(args.ddl_sql),Path(args.out_dir),ai_cfg)
    except Exception: traceback.print_exc()

if __name__=="__main__": main()
