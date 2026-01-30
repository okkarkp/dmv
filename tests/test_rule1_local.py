from openpyxl import load_workbook

def main():
    wb = load_workbook("tests/data/rule1_out.xlsx", read_only=True)
    ws = wb["Baseline Data Model_output"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    r1 = headers.index("Rule1")
    status = headers.index("Validation_Status")
    mig = headers.index("Migrating or Not (Yes/No)")
    reason = headers.index("Reason for Not Migrating")

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 1) All migrating = YES must PASS
    for r in rows:
        if str(r[mig]).upper() == "YES":
            assert r[r1] == "PASS"
            assert r[status] == "PASS"

    # 2) Migrating = NO with reason must PASS
    assert any(
        str(r[mig]).upper() == "NO"
        and r[reason] not in (None, "")
        and r[r1] == "PASS"
        for r in rows
    )

    # 3) Migrating = NO without reason must FAIL
    assert any(
        str(r[mig]).upper() == "NO"
        and r[reason] in (None, "")
        and r[r1] == "FAIL"
        for r in rows
    )

    print("[OK] Rule1 assertions passed")

if __name__ == "__main__":
    main()
