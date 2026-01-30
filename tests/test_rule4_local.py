from openpyxl import load_workbook

def main():
    wb = load_workbook("tests/data/rule4_out.xlsx", read_only=True, data_only=True)

    # 1) Check mismatch sheet contains expected issues
    ws_m = wb["Rule4_DDL_Mismatch"]
    issues = []
    for r in ws_m.iter_rows(min_row=2, values_only=True):
        issues.append(r)

    # Expect:
    # - DMW_ONLY for EXTRA_COL
    # - MISMATCH for CREATED_AT
    # - MISSING_IN_DMW for NAME (exists in DDL but not mapped in DMW)
    assert any(x[1] == "EXTRA_COL" and x[2] == "DMW_ONLY" for x in issues)
    assert any(x[1] == "CREATED_AT" and x[2] == "MISMATCH" for x in issues)
    assert any(x[1] == "NAME" and x[2] == "MISSING_IN_DMW" for x in issues)

    # 2) Baseline propagation: ID=PASS, EXTRA_COL=FAIL, CREATED_AT=FAIL
    ws_b = wb["Baseline Data Model_output"]
    header = [c.value for c in next(ws_b.iter_rows(min_row=1, max_row=1))]
    dt = header.index("Destination Table")
    dc = header.index("Destination Column Name")
    r4 = header.index("Rule4")
    st = header.index("Validation_Status")

    found = {}
    for r in ws_b.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        key = (str(r[dt]).upper(), str(r[dc]).upper())
        found[key] = (r[r4], r[st])

    assert found[("T_RULE4", "ID")] == ("PASS", "PASS")
    assert found[("T_RULE4", "EXTRA_COL")] == ("FAIL", "FAIL")
    assert found[("T_RULE4", "CREATED_AT")] == ("FAIL", "FAIL")

    print("[OK] Rule4 assertions passed")

if __name__ == "__main__":
    main()
