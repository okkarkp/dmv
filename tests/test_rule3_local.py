from openpyxl import load_workbook

def main():
    wb = load_workbook("tests/data/rule3_dmw.xlsx", read_only=True)
    ws_tbl = wb["Table Details"]

    base_tables = {"T_OK"}
    results = {}

    for r in ws_tbl.iter_rows(min_row=2, values_only=True):
        t = str(r[0]).strip().upper() if r[0] else ""
        if not t or t == "NA":
            continue
        results[t] = "PASS" if t in base_tables else "FAIL"

    assert results["T_OK"] == "PASS"
    assert results["T_MISSING"] == "FAIL"

    print("[OK] Rule3 assertions passed")

if __name__ == "__main__":
    main()
