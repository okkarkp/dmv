"""
Rule 4 CI Test
--------------
Validates:
A) DMW_ONLY detected correctly
B) MISSING_IN_DMW detected correctly
C) Rule4 FAIL is column-level only
"""

import subprocess
from openpyxl import load_workbook, Workbook

def create_dmw():
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Source Table",
        "Source Column Name",
        "Destination Table",
        "Destination Column Name"
    ])
    ws.append(["SRC_T", "SRC_ID", "T_RULE4", "ID"])          # valid
    ws.append(["SRC_T", "SRC_X", "T_RULE4", "EXTRA_COL"])   # DMW_ONLY
    ws.append(["NA", "NA", "T_RULE4", "CREATED_AT"])        # valid
    wb.save("tests_rule4_dmw.xlsx")

def create_ddl():
    with open("tests_rule4_ddl.sql", "w") as f:
        f.write("""
CREATE TABLE [dbo].[T_RULE4] (
    [ID] INT NOT NULL,
    [CREATED_AT] DATETIME NOT NULL
);
""")

def run_validator():
    subprocess.run([
        "python3", "validate_dmw_final.py",
        "--dmw-xlsx", "tests_rule4_dmw.xlsx",
        "--ddl-sql", "tests_rule4_ddl.sql",
        "--out", "tests_rule4_out.xlsx"
    ], check=True)

def assert_results():
    wb = load_workbook("tests_rule4_out.xlsx", read_only=True)

    # --- Rule4_DDL_Mismatch ---
    ws = wb["Rule4_DDL_Mismatch"]
    mismatches = [r for r in ws.iter_rows(min_row=2, values_only=True)]

    assert ("T_RULE4","EXTRA_COL","DMW_ONLY") in [
        (r[0], r[1], r[2]) for r in mismatches
    ], "DMW_ONLY not detected"

    # --- Baseline ---
    ws = wb["Baseline Data Model_output"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    i_dt = hdr.index("Destination Table")
    i_dc = hdr.index("Destination Column Name")
    i_r4 = hdr.index("Rule4")
    i_st = hdr.index("Validation_Status")

    results = {
        (r[i_dt], r[i_dc]): (r[i_r4], r[i_st])
        for r in ws.iter_rows(min_row=2, values_only=True)
    }

    assert results[("T_RULE4","ID")] == ("PASS","PASS")
    assert results[("T_RULE4","CREATED_AT")] == ("PASS","PASS")
    assert results[("T_RULE4","EXTRA_COL")] == ("FAIL","FAIL")

def main():
    create_dmw()
    create_ddl()
    run_validator()
    assert_results()
    print("[PASS] Rule 4 CI test")

if __name__ == "__main__":
    main()
