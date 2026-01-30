from openpyxl import load_workbook

def main():
    wb = load_workbook("tests/data/rule2_out.xlsx", read_only=True)
    ws = wb["Baseline Data Model_output"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    r2 = headers.index("Rule2")
    intro = headers.index("Introduced Sprint (for data migration sprint)")
    updated = headers.index("Last Updated in Sprint/Pass")
    changelog = headers.index("Chang Log (for data migration reference)")

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Case 1: Same sprint → PASS
    assert any(
        r[intro] == r[updated] and r[r2] == "PASS"
        for r in rows
    ), "Rule2 failed: same sprint should PASS"

    # Case 2: Different sprint + changelog → PASS
    assert any(
        r[intro] != r[updated]
        and r[changelog] not in (None, "")
        and r[r2] == "PASS"
        for r in rows
    ), "Rule2 failed: change log present should PASS"

    # Case 3: Different sprint + NO changelog → FAIL
    assert any(
        r[intro] != r[updated]
        and r[changelog] in (None, "")
        and r[r2] == "FAIL"
        for r in rows
    ), "Rule2 failed: missing change log should FAIL"

    print("[OK] Rule2 assertions passed")

if __name__ == "__main__":
    main()
